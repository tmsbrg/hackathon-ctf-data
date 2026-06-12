#!/usr/bin/env python3
"""Generate company flavor: cruft, duplicates, lore, PDFs, xlsx, email threads."""
from __future__ import annotations

import os
import random
import shutil
import subprocess
import textwrap
from pathlib import Path

from openpyxl import Workbook

DUMP = Path(os.environ["DUMP"])

EMPLOYEES = [
    ("Jan de Vries", "Operations", "j.devries@nordwind-logistics.nl"),
    ("Sophie Bakker", "Finance", "s.bakker@nordwind-logistics.nl"),
    ("Pieter Jansen", "Warehouse", "p.jansen@nordwind-logistics.nl"),
    ("Anna Smit", "Marketing", "a.smit@nordwind-logistics.nl"),
    ("M. van den Berg", "Executive", "m.vandenberg@nordwind-logistics.nl"),
    ("L. Chen", "Finance", "l.chen@nordwind-logistics.nl"),
    ("Erik Mulder", "Operations (contractor)", "e.mulder@nordwind-logistics.nl"),
    ("Li Wei", "Shanghai WMS", "li.wei@nordwind-logistics.cn"),
    ("Daan Jansen", "IT", "d.jansen@nordwind-logistics.nl"),
    ("Eva de Groot", "Legal / Board sec.", "e.degroot@nordwind-logistics.nl"),
]

CRUFT_DIRS = [
    "Shared/meeting_notes",
    "Marketing/brand_assets",
    "HR/onboarding",
    "Finance/invoices/2024",
    "IT/onboarding",
    "Operations/warehouse",
    "Legal/contracts",
    "China_Office/training",
    "Shared/templates",
    "Archives/2020",
]

POLICY_PDFS = [
    (
        "HR/policies/reiskostenvergoeding_2024.pdf",
        "nl",
        """\
        # Reiskostenvergoeding 2024

        **Nordwind Logistics BV** — HR-POL-008

        Medewerkers ontvangen EUR 0,23 per km woon-werkverkeer (max 75 km enkele reis).
        Declaraties via AFAS uiterlijk de 5e van de maand.

        Contact: HR@nordwind-logistics.nl
        """,
    ),
    (
        "HR/policies/travel_expense_policy_2024.pdf",
        "en",
        """\
        # Travel Expense Policy 2024

        **Nordwind Logistics BV** — HR-POL-008-EN

        Business travel must be pre-approved by line manager.
        Economy class for flights under 6 hours. Submit receipts within 30 days.

        Contact: HR@nordwind-logistics.nl
        """,
    ),
    (
        "Operations/customs/Brexit_border_checklist_NL.pdf",
        "nl",
        """\
        # Brexit Grenscontrole Checklist

        **Nordwind Logistics BV** — OPS-FORM-044

        Controleer voor elke UK-zending: T1 document, HS-code, Incoterms, EORI-nummer klant.
        Bij twijfel: douane-afdeling ext. 4422.

        Versie 2.1 — geldig vanaf 1 januari 2024.
        """,
    ),
    (
        "HR/policies/AVG_medewerker_informatie.pdf",
        "nl",
        """\
        # Informatie over gegevensverwerking (AVG)

        **Nordwind Logistics BV**

        Wij verwerken persoonsgegevens voor salarisadministratie, planning en veiligheid.
        Functionaris gegevensbescherming: privacy@nordwind-logistics.nl

        U heeft recht op inzage, correctie en verwijdering conform AVG art. 15–17.
        """,
    ),
    (
        "Operations/warehouse/health_safety_warehouse.pdf",
        "en",
        """\
        # Warehouse Health & Safety Briefing

        **Nordwind Logistics BV** — Rotterdam DC

        PPE mandatory in zones B and C. Forklift lanes: pedestrians use marked crossings.
        Report incidents to shift lead within 15 minutes.

        Emergency assembly point: parking lot P2 north.
        """,
    ),
]


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def mkpdf_from_md(md_content: str, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    import tempfile

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".md", delete=False, encoding="utf-8"
    ) as f:
        f.write(textwrap.dedent(md_content).strip() + "\n")
        md_tmp = f.name
    try:
        subprocess.run(["pandoc", md_tmp, "-o", str(out_path)], check=True)
    finally:
        Path(md_tmp).unlink(missing_ok=True)


def cruft_pass() -> None:
    for rel in CRUFT_DIRS:
        folder = DUMP / rel
        folder.mkdir(parents=True, exist_ok=True)

        write_text(
            folder / "desktop.ini",
            """\
            [.ShellClassInfo]
            ConfirmFileOp=0
            IconResource=C:\\Windows\\system32\\shell32.dll,3
            [ViewState]
            Mode=
            Vid={{137E7700-3573-11CF-AE69-08002B2E1262}}
            FolderType=Documents
            """,
        )

        # Minimal placeholder binaries — common on real SMB shares
        (folder / "Thumbs.db").write_bytes(
            b"THUMBS\x00placeholder\x00" + os.urandom(48)
        )
        (folder / ".DS_Store").write_bytes(
            b"\x00\x00\x00\x01Bud1\x00\x00\x10\x00" + os.urandom(24)
        )

    # Word lock file — someone had VPN doc open during the mirror
    lock = DUMP / "IT/onboarding" / "~$VPN_toegang_2024.docx"
    lock.write_bytes(
        b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
        b"Microsoft Office Word lock file placeholder"
    )


def duplicate_pass() -> None:
    pairs = [
        (
            DUMP / "Finance/invoices/2024/INV-2024-0007.txt",
            DUMP / "Finance/invoices/2024/INV-2024-0007 (1).txt",
        ),
        (
            DUMP / "Shared/meeting_notes/notes_2024_012.txt",
            DUMP / "Shared/meeting_notes/notes_2024_012 (1).txt",
        ),
        (
            DUMP / "Legal/contracts/NDA_counterparty_0003.txt",
            DUMP / "Shared/templates/NDA_counterparty_0003_old.txt",
        ),
        (
            DUMP / "HR/policies/policy_HR_005.txt",
            DUMP / "Archives/2020/policy_HR_005_archived.txt",
        ),
        (
            DUMP / "Operations/routes/route_0004.json",
            DUMP / "Operations/routes/route_0004_backup.json",
        ),
    ]
    for src, dst in pairs:
        if src.exists():
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)


def lore_pass() -> None:
    lore_files = [
        (
            "Operations/warehouse/driver_roster_week42.txt",
            "Week 42 driver roster — Rotterdam hub\n"
            "Mon: Pieter Jansen (NW-00633), Tue: Jan de Vries (NW-00421)\n"
            "Wed: Pieter Jansen, Thu: Jan de Vries, Fri: split shift\n"
            "Notes: Pieter covering Jan's route Thu PM (training).",
        ),
        (
            "Finance/audit/auditor_document_request_2024.txt",
            "PwC audit — document request list\n"
            "Contact at Nordwind: Sophie Bakker (s.bakker@nordwind-logistics.nl)\n"
            "Please provide Q3 payroll export by 15 Oct. CEO sign-off: M. van den Berg.",
        ),
        (
            "IT/software_inventory.csv",
            "product,version,owner,last_patched\n"
            "GlobalProtect,6.2,d.jansen@nordwind-logistics.nl,2024-10-01\n"
            "Exact Online,cloud,s.bakker@nordwind-logistics.nl,n/a\n"
            "AFAS HR,cloud,e.degroot@nordwind-logistics.nl,n/a\n",
        ),
        (
            "Shared/meeting_notes/koffiepauze_rooster.txt",
            "Koffiepauze rooster — 3e verdieping\n"
            "Week 41: Anna Smit | Week 42: Sophie Bakker | Week 43: Eva de Groot\n"
            "Regel: filter vervangen als hij rood knippert. Geen pods in het afval.",
        ),
        (
            "Marketing/campaigns/2024/linkedin_post_draft_v2.txt",
            "Draft LinkedIn post — Sophie asked to review numbers before publish.\n"
            "Quote from M. van den Berg on Benelux growth. Anna Smit to add hashtag set.\n"
            "Status: awaiting CFO approval from L. Chen.",
        ),
        (
            "Operations/routes/incident_flat_tire_A12.txt",
            "Incident report 2024-09-18 — A12 eastbound\n"
            "Driver: Jan de Vries. Delay 45 min. Cargo intact.\n"
            "Recovery: Pieter Jansen dispatched spare truck NW-TRK-442.",
        ),
        (
            "China_Office/reports/shanghai_hq_sync_call.txt",
            "Monthly sync EU HQ <> Shanghai\n"
            "Attendees: L. Chen, Li Wei, Daan Jansen (IT bridge)\n"
            "Li Wei confirmed CN-SHA-01 throughput target met. WMS training photos uploaded.",
        ),
        (
            "HR/exit_interviews/contractor_mulder_renewal.txt",
            "Contractor renewal checklist — Erik Mulder\n"
            "Ops manager: Jan de Vries recommends extend 6 months.\n"
            "Finance: rate unchanged. HR: scan signed contract to DMS (see Mulder_contract_scan.pdf).",
        ),
        (
            "Legal/board/board_meeting_catering.txt",
            "Board meeting catering order 18 Dec 2024\n"
            "Ordered by Eva de Groot. Attendees include M. van den Berg, L. Chen.\n"
            "Vendor: Ahoy Catering. Dietary: 1 gluten-free (Sophie Bakker).",
        ),
        (
            "IT/tickets/printer_floor3_jam.txt",
            "IT Ticket #9012 — Closed\n"
            "User: Anna Smit. Subject: Printer floor 3 jam.\n"
            "Resolution: cleared tray. User asked if password for printer admin is password123 — denied.",
        ),
        (
            "Finance/fuel_card_reconciliation_sep.txt",
            "Shell Fleet card reconciliation September 2024\n"
            "Prepared by Sophie Bakker. Anomaly: Pieter Jansen card used Antwerp twice same day — verified OK.",
        ),
        (
            "Operations/warehouse/shift_handover_jansen.txt",
            "Shift handover 2024-10-02 night -> day\n"
            "Outgoing: Pieter Jansen. Incoming: Jan de Vries.\n"
            "Bay 7 conveyor belt noisy — maintenance ticket #9018 opened by Daan Jansen.",
        ),
        (
            "Marketing/brand_assets/photo_shoot_shanghai_notes.txt",
            "Brand photo shoot notes — Shanghai warehouse\n"
            "Photographer met Li Wei on site. Do not use terminal screenshots in marketing.\n"
            "Anna Smit to review selects by Friday.",
        ),
        (
            "Shared/meeting_notes/team_uitje_planning.txt",
            "Team outing planning committee\n"
            "Location shortlist: SS Rotterdam, Euromast. Budget owner: Sophie Bakker.\n"
            "Anna Smit polling Slack. Jan de Vries volunteered BBQ committee.",
        ),
        (
            "IT/projects/shanghai_wms_migration_status.txt",
            "Project: Shanghai WMS integration\n"
            "PM: Daan Jansen. Business: Li Wei. Sponsor: L. Chen.\n"
            "Status: green. Go-live CN-SHA-01 complete. Hypercare until Nov 30.",
        ),
        (
            "HR/onboarding/new_starter_checklist_template.txt",
            "New starter checklist template\n"
            "Buddy assignment, badge photo, AFAS account — owner HR shared mailbox.\n"
            "Example completed for Anna Smit (2022 intake) on file.",
        ),
        (
            "Finance/invoices/2024/INTERNAL_catering_board_meeting.txt",
            "INTERNAL memo — do not send to vendor\n"
            "Ahoy Catering invoice tied to board meeting 18 Dec.\n"
            "Approver: Eva de Groot. Budget code: EXEC-2024-044.",
        ),
        (
            "Operations/customs/client_eori_verification_log.txt",
            "EORI verification log excerpt\n"
            "Verified client EORI for Jan de Vries' route batch RT-EU-0012 — OK.\n"
            "Douane contact used checklist Brexit_border_checklist_NL.pdf.",
        ),
        (
            "China_Office/training/li_wei_safety_briefing_attendance.txt",
            "Safety briefing attendance CN-SHA-01 — August 2024\n"
            "Present: Li Wei + 34 warehouse staff. Topic: forklift/pedestrian lanes.\n"
            "English summary sent to Daan Jansen for HQ records.",
        ),
        (
            "Shared/recipes/staff_birthdays_2024.txt",
            "Staff birthdays 2024 (internal — do not circulate externally)\n"
            "Jan de Vries — 15 March | Sophie Bakker — 2 July | Pieter Jansen — 21 Nov\n"
            "Anna Smit — 8 May | Eva de Groot — 30 Jan | Daan Jansen — 14 Sep\n"
            "Card collection: Shared/meeting_notes/koffiepauze_rooster.txt volunteers.",
        ),
    ]
    for rel, body in lore_files:
        write_text(DUMP / rel, body)


def pandoc_pdf_pass() -> None:
    for rel, _lang, md in POLICY_PDFS:
        mkpdf_from_md(md, DUMP / rel)


def xlsx_pass() -> None:
    # Route schedule
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Routes"
    ws.append(["route_id", "driver", "origin", "destination", "departure"])
    ws.append(["RT-EU-0012", "Jan de Vries", "Rotterdam", "Antwerp", "06:00"])
    ws.append(["RT-EU-0042", "Pieter Jansen", "Rotterdam", "Hamburg", "07:30"])
    ws.append(["RT-EU-0088", "Jan de Vries", "Rotterdam", "Paris", "05:45"])
    p = DUMP / "Operations/routes/RT-EU-schedule_2024.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)

    # Budget draft
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget 2025"
    ws.append(["Department", "Q1", "Q2", "Q3", "Q4", "Owner"])
    ws.append(["Operations", 420000, 430000, 440000, 450000, "Jan de Vries"])
    ws.append(["Finance", 180000, 185000, 190000, 195000, "Sophie Bakker"])
    ws.append(["Marketing", 95000, 120000, 110000, 130000, "Anna Smit"])
    ws.append(["IT", 210000, 215000, 220000, 225000, "Daan Jansen"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["NOTE", "DRAFT — not approved", "CFO review pending", "", "", "L. Chen"])
    p = DUMP / "Finance/budget_2025_draft_v3_FINAL.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)

    # TEU throughput
    wb = Workbook()
    ws = wb.active
    ws.title = "TEU Q3"
    ws.append(["week", "teu_in", "teu_out", "shift_lead"])
    for week, teu in [(27, 1180), (28, 1220), (29, 1195), (30, 1240), (31, 1210)]:
        ws.append([week, teu, teu - random.randint(20, 80), "Pieter Jansen"])
    p = DUMP / "Operations/warehouse/teu_throughput_Q3.xlsx"
    wb.save(p)


def eml_thread_pass() -> None:
    thread = [
        (
            "Shared/meeting_notes/email_thread_office_move_01.eml",
            """\
            From: e.degroot@nordwind-logistics.nl
            To: all-staff@nordwind-logistics.nl
            Subject: Verhuizing afdeling Finance — verdieping 4
            Date: Mon, 14 Oct 2024 08:15:00 +0200
            Message-ID: <move-announce-01@nw-dc01>

            Beste collega's,

            Per 4 november verhuist Finance naar verdieping 4 (westvleugel).
            Sophie Bakker is aanspreekpunt voor crate-nummers en labelmachines.

            Printer op 3e blijft beschikbaar tot 8 nov.

            Groet,
            Eva de Groot
            Office Management
            """,
        ),
        (
            "Shared/meeting_notes/email_thread_office_move_02.eml",
            """\
            From: a.smit@nordwind-logistics.nl
            To: e.degroot@nordwind-logistics.nl
            Cc: helpdesk@nordwind-logistics.nl
            Subject: RE: Verhuizing afdeling Finance — verdieping 4
            Date: Mon, 14 Oct 2024 10:02:33 +0200
            In-Reply-To: <move-announce-01@nw-dc01>
            Message-ID: <move-reply-02@nw-dc01>

            Hi Eva,

            Marketing shares the corridor printer on 3 — can we get a queue rename
            before Finance moves? My jobs keep landing on FIN-3F-COLOR.

            Thanks,
            Anna
            """,
        ),
        (
            "Shared/meeting_notes/email_thread_office_move_03.eml",
            """\
            From: helpdesk@nordwind-logistics.nl
            To: a.smit@nordwind-logistics.nl
            Cc: e.degroot@nordwind-logistics.nl; d.jansen@nordwind-logistics.nl
            Subject: RE: Verhuizing afdeling Finance — verdieping 4
            Date: Mon, 14 Oct 2024 14:47:09 +0200
            In-Reply-To: <move-reply-02@nw-dc01>
            Message-ID: <move-reply-03@nw-dc01>

            Hi Anna,

            Renamed queue to MKT-SHARED-3F. Old FIN-3F-COLOR redirects until 8 Nov.
            If jam error persists, power-cycle — do not use admin PIN from sticky notes.

            Ticket #9012 closed.

            IT Helpdesk
            """,
        ),
    ]
    for rel, body in thread:
        write_text(DUMP / rel, body)


def main() -> None:
    random.seed(42)
    cruft_pass()
    lore_pass()
    pandoc_pdf_pass()
    xlsx_pass()
    eml_thread_pass()
    duplicate_pass()  # after lore creates some source files
    print(
        "Flavor pass complete (cruft, lore, PDFs, xlsx, email thread, duplicates)."
    )


if __name__ == "__main__":
    main()
