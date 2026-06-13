#!/usr/bin/env python3
"""Generate legacy Office formats via LibreOffice CLI (.doc, .xlsx, .xls).

rga typically cannot search these — participants use unzip (xlsx) or
LibreOffice/strings (doc/xls). One xlsx gets a hidden sheet via openpyxl
post-processing after LibreOffice creates the base file.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import textwrap
from pathlib import Path

DUMP = Path(os.environ["DUMP"])

LO_BIN = shutil.which("libreoffice") or shutil.which("soffice")
DOC_FILTER = "doc:MS Word 2007 XML"
XLS_FILTER = "xls:MS Excel 97"

SAP_VENDOR_API = "sap_api_NwVendor_8842secret"
LEGACY_HR_PWD = "LegacyHr-Wachtwoord-8842!"
EMERGENCY_TOKEN = "BgNw-Emergency-2024-xK9"
FLEET_API = "fleet_api_NwShell_8842secret"
XLS_ARCHIVE_KEY = "ArchNw-Xls-8842legacy"


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def lo_env(profile: Path) -> dict[str, str]:
    env = os.environ.copy()
    env["HOME"] = str(profile)
    return env


def lo_convert(profile: Path, src: Path, outdir: Path, convert_to: str) -> Path:
    if not LO_BIN:
        raise RuntimeError(
            "libreoffice/soffice not found — install LibreOffice for .doc/.xlsx generation"
        )
    outdir.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            LO_BIN,
            "--headless",
            "--norestore",
            "--convert-to",
            convert_to,
            "--outdir",
            str(outdir),
            str(src),
        ],
        check=True,
        capture_output=True,
        env=lo_env(profile),
    )
    # LibreOffice keeps the source stem; find newest matching output
    suffix = convert_to.split(":")[0].split('"')[0]
    candidates = sorted(outdir.glob(f"{src.stem}.{suffix}*"), key=lambda p: p.stat().st_mtime)
    if not candidates:
        raise RuntimeError(f"LibreOffice did not produce {suffix} for {src}")
    return candidates[-1]


def html_to_doc(profile: Path, work: Path, html_body: str, dest: Path) -> None:
    html_path = work / "source.html"
    html_path.write_text(
        f"<!DOCTYPE html><html><head><meta charset='utf-8'></head><body>{html_body}</body></html>",
        encoding="utf-8",
    )
    produced = lo_convert(profile, html_path, work, DOC_FILTER)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(produced), str(dest))


def csv_to_xlsx(profile: Path, work: Path, csv_content: str, dest: Path) -> None:
    csv_path = work / "source.csv"
    csv_path.write_text(textwrap.dedent(csv_content).strip() + "\n", encoding="utf-8")
    produced = lo_convert(profile, csv_path, work, "xlsx")
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(produced), str(dest))


def csv_to_xls(profile: Path, work: Path, csv_content: str, dest: Path) -> None:
    csv_path = work / "source.csv"
    csv_path.write_text(textwrap.dedent(csv_content).strip() + "\n", encoding="utf-8")
    produced = lo_convert(profile, csv_path, work, XLS_FILTER)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(produced), str(dest))


def add_hidden_sheet(xlsx_path: Path, sheet_name: str, rows: list[list[str]]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    hidden = wb.create_sheet(sheet_name)
    hidden.sheet_state = "hidden"
    for row in rows:
        hidden.append(row)
    wb.save(xlsx_path)


def generate_docs(profile: Path, work: Path) -> None:
    html_to_doc(
        profile,
        work,
        f"""\
        <h1>SAP Vendor Integration — Internal Memo</h1>
        <p><strong>Nordwind Logistics BV</strong> — Finance / Procurement</p>
        <p>Date: 14 March 2024 | Author: Sophie Bakker</p>
        <p>SAP B1 connector credentials for PostNL freight reconciliation (sandbox migrated to prod):</p>
        <table border="1" cellpadding="4">
          <tr><td>Endpoint</td><td>https://sap.nordwind.internal/api/v1</td></tr>
          <tr><td>Integration user</td><td>sap_integration</td></tr>
          <tr><td>API key</td><td><strong>{SAP_VENDOR_API}</strong></td></tr>
        </table>
        <p>Rotate quarterly. Do not email this memo.</p>
        """,
        DUMP / "Finance/invoices/2024/SAP_vendor_integration_memo.doc",
    )

    html_to_doc(
        profile,
        work,
        f"""\
        <h1>IT Account Reset Procedure (Legacy)</h1>
        <p><strong>Nordwind Logistics BV</strong> — HR / IT joint procedure 2019</p>
        <p>Archived on SMB — superseded by SSO 2022.</p>
        <h2>Break-glass HR admin account</h2>
        <p>Used when AFAS sync fails and director cannot access HR portal.</p>
        <ul>
          <li>Account: <code>hr_breakglass</code></li>
          <li>Wachtwoord (legacy): <strong>{LEGACY_HR_PWD}</strong></li>
          <li>Ticket reference: IT-4412</li>
        </ul>
        <p>Contact helpdesk before use. Document in ticket.</p>
        """,
        DUMP / "HR/policies/IT_account_reset_procedure_2019.doc",
    )

    html_to_doc(
        profile,
        work,
        """\
        <h1>Office Etiquette 2017</h1>
        <p>Nordwind Logistics BV — HR circular</p>
        <ul>
          <li>Clean microwave after use.</li>
          <li>No fish in open plan on Fridays.</li>
          <li>Label food in shared fridge with name and date.</li>
        </ul>
        <p>No confidential information in this document.</p>
        """,
        DUMP / "HR/policies/office_etiquette_2017.doc",
    )


def generate_spreadsheets(profile: Path, work: Path) -> None:
    emergency_xlsx = DUMP / "Finance/payroll/emergency_access_roster_Q4.xlsx"
    csv_to_xlsx(
        profile,
        work,
        f"""\
        role,employee_id,name,department,emergency_token,notes
        CFO backup,NW-00887,Sophie Bakker,Finance,,
        CEO portal,NW-EXEC01,M. van den Berg,Executive,{EMERGENCY_TOKEN},rotate after tabletop
        IT break-glass,NW-IT99,hr_breakglass,IT,see_legacy_doc,AFAS outage only
        Warehouse lead,NW-00633,Pieter Jansen,Warehouse,,on-call weekends
        """,
        emergency_xlsx,
    )

    fuel_xlsx = DUMP / "Operations/routes/fuel_card_registry_2024.xlsx"
    csv_to_xlsx(
        profile,
        work,
        """\
        card_id,driver,vehicle_reg,provider,status
        FC-8842,Jan de Vries,NW-TRK-112,Shell Fleet,active
        FC-8843,Pieter Jansen,NW-TRK-442,Shell Fleet,active
        FC-9012,Anna Smit,NW-VAN-088,Shell Fleet,active
        FC-7701,Li Wei (EU visit),NW-TRK-901,Shell Fleet,suspended
        """,
        fuel_xlsx,
    )
    add_hidden_sheet(
        fuel_xlsx,
        "shell_api",
        [
            ["Shell Fleet API integration (do not share)"],
            ["api_key", FLEET_API],
            ["endpoint", "https://fleet.shell.com/api/v2"],
            ["owner", "s.bakker@nordwind-logistics.nl"],
        ],
    )

    csv_to_xls(
        profile,
        work,
        f"""\
        invoice_id,year,vendor,amount_eur,archive_export_key
        INV-2023-0892,2023,PostNL,1240.50,
        INV-2023-0893,2023,Shell Fleet,3890.00,
        INV-2023-0894,2023,KPN Business,890.25,{XLS_ARCHIVE_KEY}
        INV-2023-0895,2023,Office Depot,445.00,
        """,
        DUMP / "Finance/invoices/2023/archive_invoice_index_2023.xls",
    )


def main() -> None:
    if not LO_BIN:
        raise SystemExit("libreoffice/soffice required — install LibreOffice")

    DUMP.mkdir(parents=True, exist_ok=True)
    profile = Path(tempfile.mkdtemp(prefix="nw-lo-profile-"))
    work = Path(tempfile.mkdtemp(prefix="nw-lo-work-"))

    try:
        generate_docs(profile, work)
        generate_spreadsheets(profile, work)
    finally:
        shutil.rmtree(profile, ignore_errors=True)
        shutil.rmtree(work, ignore_errors=True)

    print(
        "Office pass complete (LibreOffice .doc, .xlsx, .xls — rga-unfriendly formats)."
    )


if __name__ == "__main__":
    main()
